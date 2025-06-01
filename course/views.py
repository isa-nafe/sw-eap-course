from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.urls import reverse, reverse_lazy
from openpyxl import load_workbook
import os
import stripe # Stripe library
from django.conf import settings
from .models import CourseMaterial, UserPayment
from django.http import HttpResponseForbidden, FileResponse, JsonResponse, HttpResponse # Added JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt # For Stripe webhook, though not fully implementing webhook here
from django.utils import timezone # For payment_date

# Initialize Stripe with secret key
stripe.api_key = settings.STRIPE_SECRET_KEY

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            try:
                students_group = Group.objects.get(name='Students')
                user.groups.add(students_group)
            except Group.DoesNotExist:
                print("Warning: 'Students' group not found.")
            login(request, user)
            # UserPayment profile is created by signal
            return redirect('course:dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'course/register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                next_url = request.GET.get('next', reverse_lazy('course:dashboard'))
                return redirect(next_url)
    else:
        form = AuthenticationForm()
    return render(request, 'course/login.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('course:login')

@login_required
def dashboard_view(request):
    is_instructor = request.user.groups.filter(name='Instructors').exists()
    user_payment, created = UserPayment.objects.get_or_create(user=request.user) # Ensure profile exists
    has_paid = user_payment.has_paid_for_course
    context = {
        'is_instructor': is_instructor,
        'has_paid': has_paid,
        'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY, # For client-side Stripe.js if needed
        'course_price_dollars': settings.COURSE_PRICE_DOLLARS,
    }
    return render(request, 'course/dashboard.html', context)

@login_required
def course_planner_view(request):
    # Payment check for planner (optional, can be free preview)
    # user_payment = get_object_or_404(UserPayment, user=request.user)
    # if not user_payment.has_paid_for_course and not request.user.is_staff:
    #     return redirect('course:payment_required') # Create a payment_required page or redirect to buy

    planner_data = []
    headers = []
    file_path = os.path.join(settings.BASE_DIR, 'course', 'data', 'planner.xlsx')
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        header_cells = sheet[1]
        headers = [cell.value for cell in header_cells]
        for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2)):
            if not any(cell.value for cell in row_cells): break
            row_data = {headers[col_idx]: cell.value for col_idx, cell in enumerate(row_cells) if col_idx < len(headers)}
            planner_data.append(row_data)
    except FileNotFoundError: print(f"Planner file not found at {file_path}")
    except Exception as e: print(f"Error reading planner file: {e}")
    return render(request, 'course/course_planner.html', {'planner_data': planner_data, 'headers': headers})

@login_required
def course_materials_list_view(request):
    user_payment = get_object_or_404(UserPayment, user=request.user)
    if not user_payment.has_paid_for_course and not request.user.is_staff and not request.user.groups.filter(name='Instructors').exists():
        # You can redirect to a "please pay" page or show a message
        # return redirect('course:dashboard') # Or a specific "enroll" page
        return render(request, 'course/course_materials_list.html', {'materials': None, 'payment_required': True, 'course_price_dollars': settings.COURSE_PRICE_DOLLARS})


    materials = CourseMaterial.objects.all().order_by('material_type', 'uploaded_at')
    return render(request, 'course/course_materials_list.html', {'materials': materials, 'payment_required': False})

@login_required
def serve_protected_material(request, material_id):
    user_payment = get_object_or_404(UserPayment, user=request.user)
    if not user_payment.has_paid_for_course and not request.user.is_staff and not request.user.groups.filter(name='Instructors').exists():
        return HttpResponseForbidden("Access Denied: Payment required for this course material.")

    material = get_object_or_404(CourseMaterial, pk=material_id)
    file_path = os.path.join(settings.MEDIA_ROOT, str(material.file))
    if not os.path.exists(file_path):
        print(f"ERROR: File not found for material ID {material_id} at path {file_path}")
        return HttpResponseForbidden("File not found on server.")
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(material.file.name))

@login_required
def create_checkout_session(request):
    YOUR_DOMAIN = request.build_absolute_uri('/')[:-1] # http://localhost:8000 or your domain

    try:
        # Create a Stripe Customer for the user if one doesn't exist
        user_payment, _ = UserPayment.objects.get_or_create(user=request.user)
        if not user_payment.stripe_customer_id:
            customer = stripe.Customer.create(
                email=request.user.email,
                name=request.user.username,
                metadata={'user_id': request.user.id}
            )
            user_payment.stripe_customer_id = customer.id
            user_payment.save()

        checkout_session = stripe.checkout.Session.create(
            customer=user_payment.stripe_customer_id,
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'usd',
                    'product_data': {
                        'name': 'Full Course Access',
                        'description': 'Lifetime access to all course materials and future updates.',
                    },
                    'unit_amount': settings.COURSE_PRICE_CENTS, # Amount in cents
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url=YOUR_DOMAIN + reverse('course:payment_success') + '?session_id={CHECKOUT_SESSION_ID}',
            cancel_url=YOUR_DOMAIN + reverse('course:payment_cancelled'),
            metadata={'user_id': request.user.id} # Pass user_id to identify user in webhook or success page
        )
        return redirect(checkout_session.url, code=303)
    except Exception as e:
        print(f"Error creating Stripe session: {e}")
        # Potentially add a user-facing error message via Django messages framework
        # messages.error(request, "Could not connect to payment gateway. Please try again later.")
        return redirect('course:dashboard') # Redirect to a safe page

@login_required
def payment_success_view(request):
    session_id = request.GET.get('session_id')
    if not session_id:
        # Handle error: no session_id provided
        return redirect('course:dashboard') # Or an error page

    try:
        session = stripe.checkout.Session.retrieve(session_id)
        # Verify session status if needed, e.g., session.payment_status == 'paid'

        if session.payment_status == 'paid':
            user_payment = get_object_or_404(UserPayment, user=request.user) # Or lookup by customer ID from session

            # Ensure this user matches the session (e.g. by user_id in metadata or customer_id)
            if session.metadata.get('user_id') != str(request.user.id) and session.customer != user_payment.stripe_customer_id:
                 # Mismatch! Handle this security concern.
                 return HttpResponseForbidden("Session user mismatch.")

            user_payment.has_paid_for_course = True
            user_payment.stripe_payment_intent_id = session.payment_intent # Store payment intent ID
            user_payment.payment_date = timezone.now()
            user_payment.amount_paid = session.amount_total / 100.0 # Convert cents to dollars
            user_payment.save()

            # Add a success message for the user
            # messages.success(request, "Payment successful! You now have full access to the course.")
        else:
            # Handle cases where payment_status is not 'paid' (e.g. 'unpaid', 'no_payment_required')
            # messages.warning(request, "Payment status is not 'paid'. Please contact support if you believe this is an error.")
            pass

    except stripe.error.StripeError as e:
        print(f"Stripe error retrieving session: {e}")
        # messages.error(request, "Error verifying payment. Please contact support.")
        return redirect('course:dashboard')
    except UserPayment.DoesNotExist:
        print(f"UserPayment profile not found for user {request.user.id} during success handling.")
        # messages.error(request, "Your payment profile was not found. Please contact support.")
        return redirect('course:dashboard')


    return render(request, 'course/payment_success.html')

@login_required
def payment_cancelled_view(request):
    # messages.info(request, "Your payment process was cancelled. You can try again anytime from your dashboard.")
    return render(request, 'course/payment_cancelled.html')

# Basic (non-CSRF protected) webhook handler for Stripe - NOT FULLY IMPLEMENTED FOR SIMPLICITY
# For production, this needs to be robust, CSRF protected if not using specific Stripe webhook verification,
# and should handle various event types.
@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
        )
    except ValueError as e: # Invalid payload
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e: # Invalid signature
        return HttpResponse(status=400)

    # Handle the checkout.session.completed event
    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        user_id = session.get('metadata', {}).get('user_id')
        stripe_customer_id = session.get('customer')

        if session.payment_status == "paid":
            try:
                user_payment = None
                if user_id:
                    user = User.objects.get(id=user_id)
                    user_payment = UserPayment.objects.get(user=user)
                elif stripe_customer_id: # Fallback if user_id not in metadata
                    user_payment = UserPayment.objects.get(stripe_customer_id=stripe_customer_id)

                if user_payment and not user_payment.has_paid_for_course: # Process only if not already marked paid
                    user_payment.has_paid_for_course = True
                    user_payment.stripe_payment_intent_id = session.payment_intent
                    user_payment.payment_date = timezone.now()
                    user_payment.amount_paid = session.amount_total / 100.0
                    user_payment.save()
                    print(f"Webhook: Payment successful for user {user_payment.user.username}")
            except User.DoesNotExist:
                print(f"Webhook Error: User with id {user_id} not found.")
            except UserPayment.DoesNotExist:
                print(f"Webhook Error: UserPayment profile not found for user_id {user_id} or customer_id {stripe_customer_id}.")
            except Exception as e:
                print(f"Webhook handling error: {e}")
    # ... handle other event types
    return HttpResponse(status=200)
